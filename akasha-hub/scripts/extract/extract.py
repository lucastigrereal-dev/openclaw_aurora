#!/usr/bin/env python3
"""
Akasha Content Extractor + Niche Classifier
Multi-format: PDF, Video, Audio, Text, Image (OCR)
Classification: GPT-4o-mini (cheap model)

Schema real:
  files: id, file_name, file_ext, mime_type, file_size_bytes, status, gdrive_id, ...
  file_content: id, file_id, content_type, content, word_count, extraction_method, extraction_cost, created_at
"""

import os
import json
import base64
import subprocess
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Tuple, Optional

from dotenv import load_dotenv
_env_path = Path(__file__).resolve().parents[3] / '.env'
if _env_path.exists():
    load_dotenv(_env_path)

from supabase import create_client, Client
from openai import OpenAI

# Supabase
SUPABASE_URL = os.getenv("AKASHA_SUPABASE_URL")
SUPABASE_KEY = os.getenv("AKASHA_SUPABASE_KEY")
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# OpenAI
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

# MIME type to extraction type mapping
MIME_TO_TYPE = {
    'application/pdf': 'pdf',
    'video/mp4': 'video', 'video/quicktime': 'video', 'video/x-msvideo': 'video',
    'video/x-matroska': 'video', 'video/webm': 'video',
    'audio/mpeg': 'audio', 'audio/wav': 'audio', 'audio/x-m4a': 'audio',
    'audio/ogg': 'audio', 'audio/flac': 'audio',
    'text/plain': 'text', 'text/markdown': 'text',
    'text/csv': 'spreadsheet',
    'application/json': 'text',
    'image/jpeg': 'image', 'image/png': 'image', 'image/gif': 'image',
    'image/webp': 'image',
    'application/vnd.openxmlformats-officedocument.wordprocessingml.document': 'document',
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': 'spreadsheet',
    'application/vnd.openxmlformats-officedocument.presentationml.presentation': 'presentation',
    'application/vnd.google-apps.document': 'document',
    'application/vnd.google-apps.spreadsheet': 'spreadsheet',
}

# Which MIME patterns to filter for each type
MIME_FILTERS = {
    'pdf': ['application/pdf'],
    'video': ['video/mp4', 'video/quicktime', 'video/x-msvideo', 'video/webm'],
    'audio': ['audio/'],
    'text': ['text/plain'],
    'markdown': ['text/markdown'],
    'json': ['application/json'],
    'image': ['image/jpeg', 'image/png'],
    'document': ['application/vnd.openxmlformats-officedocument.wordprocessingml', 'application/vnd.google-apps.document'],
}


def get_extract_type(mime_type: str) -> str:
    """Map MIME type to extraction type."""
    if not mime_type:
        return 'other'
    return MIME_TO_TYPE.get(mime_type, 'other')


class ContentExtractor:
    """Multi-format content extraction + niche classification"""

    def __init__(self):
        self.temp_dir = Path.home() / ".openclaw/hubs/akasha/temp"
        self.temp_dir.mkdir(parents=True, exist_ok=True)

    # Files that are NOT knowledge sources - skip these
    JUNK_NAMES = {
        'license', 'licence', 'changelog', 'contributing', 'readme',
        'notice', 'authors', 'makefile', 'dockerfile', 'vagrantfile',
        'gemfile', 'rakefile', 'gruntfile', 'gulpfile', 'procfile',
    }
    JUNK_PATTERNS = [
        'license', 'licence', '1000-std', 'thirdparty', 'changelog',
        'contributing', 'code_of_conduct', '.log', 'test-output',
        'node_modules', '__pycache__', '.git/', 'dist/',
    ]

    def _is_knowledge_file(self, file_record: Dict) -> bool:
        """Check if a file is likely a knowledge source (not junk)."""
        name = (file_record.get("file_name") or "").lower()
        path = (file_record.get("file_path") or "").lower()

        # Skip known junk filenames
        name_no_ext = name.rsplit('.', 1)[0] if '.' in name else name
        if name_no_ext in self.JUNK_NAMES:
            return False

        # Skip known junk patterns
        for pattern in self.JUNK_PATTERNS:
            if pattern in name or pattern in path:
                return False

        # Skip very small files (< 100 bytes) - likely empty/stubs
        size = file_record.get("file_size_bytes") or 0
        if size < 100:
            return False

        return True

    def get_pending_files(self, limit: int = 10, file_type: str = "all",
                          knowledge_only: bool = True, local_only: bool = False) -> List[Dict]:
        """Get files pending extraction, filtered to knowledge sources."""
        query = supabase.table("files") \
            .select("id, file_name, file_ext, mime_type, file_size_bytes, status, gdrive_id, file_path") \
            .eq("status", "cataloged")

        if file_type != "all":
            patterns = MIME_FILTERS.get(file_type, [])
            if patterns:
                query = query.ilike("mime_type", f"{patterns[0]}%")

        if local_only:
            query = query.ilike("file_path", "local://%")

        # Get more than needed, then filter for knowledge
        fetch_limit = limit * 3 if knowledge_only else limit
        response = query.order("file_size_bytes", desc=True).limit(fetch_limit).execute()

        if knowledge_only:
            filtered = [f for f in response.data if self._is_knowledge_file(f)]
            return filtered[:limit]
        return response.data

    def extract_pdf(self, file_path: Path) -> Tuple[str, Dict]:
        """Extract text from PDF using PyPDF2 (FREE)"""
        try:
            from PyPDF2 import PdfReader
        except ImportError:
            raise RuntimeError("PyPDF2 not installed: pip install PyPDF2")

        reader = PdfReader(str(file_path))
        text = ""
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"

        metadata = {
            "extraction_method": "PyPDF2",
            "page_count": len(reader.pages),
            "cost_usd": 0.0
        }

        return text.strip(), metadata

    def extract_text(self, file_path: Path) -> Tuple[str, Dict]:
        """Extract text from plain text files (FREE)"""
        encodings = ['utf-8', 'latin-1', 'cp1252']

        for encoding in encodings:
            try:
                text = file_path.read_text(encoding=encoding)
                return text, {"extraction_method": "direct_read", "encoding": encoding, "cost_usd": 0.0}
            except UnicodeDecodeError:
                continue

        raise RuntimeError(f"Could not decode {file_path}")

    def extract_docx(self, file_path: Path) -> Tuple[str, Dict]:
        """Extract text from DOCX using python-docx (FREE)"""
        from docx import Document
        doc = Document(str(file_path))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        text = "\n".join(paragraphs)
        return text, {"extraction_method": "python-docx", "paragraph_count": len(paragraphs), "cost_usd": 0.0}

    def extract_xlsx(self, file_path: Path) -> Tuple[str, Dict]:
        """Extract text from XLSX using openpyxl (FREE)"""
        from openpyxl import load_workbook
        wb = load_workbook(str(file_path), read_only=True, data_only=True)
        all_text = []
        total_rows = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_text.append(f"=== Sheet: {sheet_name} ===")
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(c.strip() for c in cells):
                    all_text.append(" | ".join(cells))
                    total_rows += 1
                if total_rows > 5000:
                    break
        wb.close()
        text = "\n".join(all_text)
        return text, {"extraction_method": "openpyxl", "total_rows": total_rows, "cost_usd": 0.0}

    def extract_csv(self, file_path: Path) -> Tuple[str, Dict]:
        """Extract text from CSV (FREE)"""
        import csv as csv_module
        encodings = ['utf-8', 'latin-1', 'cp1252']
        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    reader = csv_module.reader(f)
                    rows = []
                    for i, row in enumerate(reader):
                        rows.append(" | ".join(row))
                        if i > 5000:
                            break
                text = "\n".join(rows)
                return text, {"extraction_method": "csv_reader", "row_count": len(rows), "encoding": encoding, "cost_usd": 0.0}
            except (UnicodeDecodeError, csv_module.Error):
                continue
        raise RuntimeError(f"Could not decode CSV {file_path}")

    def extract_video_audio(self, file_path: Path, file_type: str,
                           speed_mode: str = "balanced") -> Tuple[str, Dict, float]:
        """Extract audio transcription using Whisper API"""
        speed_map = {
            "conservative": 1.2,
            "balanced": 1.5,
            "aggressive": 2.0
        }
        speed = speed_map.get(speed_mode, 1.5)

        audio_path = file_path
        if file_type == "video":
            audio_path = self.temp_dir / f"{file_path.stem}_audio.mp3"
            subprocess.run([
                "ffmpeg", "-i", str(file_path), "-vn", "-acodec", "libmp3lame",
                "-q:a", "4", str(audio_path), "-y"
            ], capture_output=True)

        if speed > 1.0:
            fast_path = self.temp_dir / f"{file_path.stem}_fast.mp3"
            subprocess.run([
                "ffmpeg", "-i", str(audio_path),
                "-filter:a", f"atempo={speed}",
                str(fast_path), "-y"
            ], capture_output=True)
            transcribe_path = fast_path
        else:
            transcribe_path = audio_path

        with open(transcribe_path, "rb") as f:
            transcript = client.audio.transcriptions.create(
                model="whisper-1",
                file=f,
                language="pt"
            )

        text = transcript.text

        file_size_mb = transcribe_path.stat().st_size / (1024 * 1024)
        duration_minutes = file_size_mb / 1.0
        cost = duration_minutes * 0.006

        metadata = {
            "extraction_method": "whisper-1",
            "speed_mode": speed_mode,
            "speed_factor": speed,
            "original_type": file_type,
            "cost_usd": round(cost, 4)
        }

        for temp_file in self.temp_dir.glob(f"{file_path.stem}_*"):
            try:
                temp_file.unlink()
            except OSError:
                pass

        return text, metadata, cost

    def extract_image_ocr(self, file_path: Path) -> Tuple[str, Dict]:
        """Extract text from image using GPT-4o Vision OCR"""
        with open(file_path, "rb") as f:
            image_data = base64.b64encode(f.read()).decode()

        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{
                "role": "user",
                "content": [
                    {"type": "text", "text": "Extraia TODO o texto desta imagem (OCR). Se houver pouco texto, descreva o conteudo visual relevante."},
                    {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{image_data}"}}
                ]
            }],
            max_tokens=1000
        )

        text = response.choices[0].message.content
        # Cost: ~0.01 per image with gpt-4o-mini
        metadata = {"extraction_method": "gpt-4o-mini-vision", "cost_usd": 0.01}
        return text, metadata

    def classify_niche(self, text: str, filename: str) -> Dict:
        """Classify content into monetization niches (GPT-4o-mini = cheap)"""
        sample = text[:2000]

        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "system",
                    "content": """Voce e um classificador de conteudo para monetizacao.
Analise o texto e retorne JSON com:
{
  "primary_niche": "um dos: estetica, marketing_digital, youtube, infoproduto, automacao, vendas, saude, financeiro, tecnologia, outro",
  "sub_niche": "sub-nicho especifico",
  "monetization_potential": "alto | medio | baixo",
  "money_score": 1-10,
  "utility_score": 1-10,
  "keywords": ["3-5 palavras-chave"],
  "summary": "resumo em 1 frase",
  "language": "pt-BR | en | es | outro"
}"""
                },
                {"role": "user", "content": f"Arquivo: {filename}\n\nConteudo:\n{sample}"}
            ],
            temperature=0.3,
            response_format={"type": "json_object"}
        )

        try:
            return json.loads(response.choices[0].message.content)
        except (json.JSONDecodeError, AttributeError):
            return {
                "primary_niche": "outro", "sub_niche": "",
                "monetization_potential": "baixo", "money_score": 1, "utility_score": 1,
                "keywords": [], "summary": "Classificacao falhou", "language": "?"
            }

    def download_drive_file(self, gdrive_id: str, mime_type: str = None) -> Optional[Path]:
        """Download file from Google Drive"""
        try:
            from googleapiclient.discovery import build
            from googleapiclient.http import MediaIoBaseDownload
            import pickle

            token_path = Path.home() / '.openclaw/google_drive_token.pickle'
            with open(token_path, 'rb') as f:
                creds = pickle.load(f)

            service = build('drive', 'v3', credentials=creds)

            request = service.files().get_media(fileId=gdrive_id)
            local_path = self.temp_dir / f"download_{gdrive_id}"

            with open(local_path, 'wb') as f:
                downloader = MediaIoBaseDownload(f, request)
                done = False
                while not done:
                    status, done = downloader.next_chunk()

            return local_path
        except Exception as e:
            print(f"   Download failed: {e}")
            return None

    def process_file(self, file_record: Dict, speed_mode: str = "balanced") -> Dict:
        """Process single file: download + extract + classify + save to DB"""
        file_id = file_record["id"]
        file_name = file_record["file_name"]
        mime_type = file_record.get("mime_type", "")
        extract_type = get_extract_type(mime_type)
        gdrive_id = file_record.get("gdrive_id")

        safe_name = file_name.encode('ascii', 'replace').decode('ascii')
        print(f"\n Processing: {safe_name}")
        print(f"   MIME: {mime_type} -> {extract_type} | Status: {file_record.get('status')}")

        report = {"file_id": file_id, "file_name": file_name, "status": "failed", "cost_usd": 0.0, "error": None}

        try:
            supabase.table("files").update({"status": "extracting"}).eq("id", file_id).execute()

            # Download from Drive or resolve local path
            if gdrive_id:
                local_path = self.download_drive_file(gdrive_id, mime_type)
                if not local_path:
                    raise RuntimeError("Download failed")
            else:
                raw_path = file_record.get("file_path", "")
                # Strip local:// prefix from scan_local.py paths
                if raw_path.startswith("local://"):
                    raw_path = raw_path[len("local://"):]
                local_path = Path(raw_path)
                if not local_path.exists():
                    raise RuntimeError(f"File not found: {local_path}")

            # Extract based on type
            extracted_text = ""
            metadata = {}
            cost = 0.0

            if extract_type == "pdf":
                extracted_text, metadata = self.extract_pdf(local_path)
            elif extract_type in ["video", "audio"]:
                extracted_text, metadata, cost = self.extract_video_audio(local_path, extract_type, speed_mode)
            elif extract_type == "document":
                # DOCX files
                ext = local_path.suffix.lower()
                if ext == '.docx':
                    extracted_text, metadata = self.extract_docx(local_path)
                else:
                    extracted_text, metadata = self.extract_text(local_path)
            elif extract_type == "spreadsheet":
                ext = local_path.suffix.lower()
                if ext == '.xlsx':
                    extracted_text, metadata = self.extract_xlsx(local_path)
                elif ext == '.csv':
                    extracted_text, metadata = self.extract_csv(local_path)
                else:
                    extracted_text, metadata = self.extract_text(local_path)
            elif extract_type == "text":
                ext = local_path.suffix.lower()
                if ext == '.csv':
                    extracted_text, metadata = self.extract_csv(local_path)
                else:
                    extracted_text, metadata = self.extract_text(local_path)
            elif extract_type == "image":
                extracted_text, metadata = self.extract_image_ocr(local_path)
                cost = metadata.get("cost_usd", 0.01)
            else:
                # Try as text fallback
                try:
                    extracted_text, metadata = self.extract_text(local_path)
                except Exception:
                    raise RuntimeError(f"Unsupported type: {extract_type} (MIME: {mime_type})")

            # Classify niche
            classification = {}
            if extracted_text and len(extracted_text.strip()) > 50:
                classification = self.classify_niche(extracted_text, file_name)
                # GPT-4o-mini classify cost ~$0.0003
                cost += 0.0003

            word_count = len(extracted_text.split()) if extracted_text else 0

            # Clean null bytes that PostgreSQL can't store
            extracted_text = extracted_text.replace('\x00', '')

            # Save to file_content (real schema)
            content_data = {
                "file_id": file_id,
                "content_type": extract_type,
                "content": extracted_text[:100000],  # limit to 100K chars
                "word_count": word_count,
                "extraction_method": metadata.get("extraction_method", "unknown"),
                "extraction_cost": round(cost, 6),
            }
            supabase.table("file_content").insert(content_data).execute()

            # Update files table with classification + status
            update_data = {
                "status": "extracted",
                "processed_at": datetime.now().isoformat(),
            }
            if classification:
                update_data["niche"] = classification.get("primary_niche", "outro")
                update_data["sub_niche"] = classification.get("sub_niche", "")
                update_data["language"] = classification.get("language", "?")
                update_data["description"] = classification.get("summary", "")
                update_data["money_score"] = classification.get("money_score", 1)
                update_data["utility_score"] = classification.get("utility_score", 1)
                tags = classification.get("keywords", [])
                if tags:
                    update_data["tags"] = tags

            supabase.table("files").update(update_data).eq("id", file_id).execute()

            report["status"] = "success"
            report["cost_usd"] = round(cost, 6)
            report["niche"] = classification.get("primary_niche", "?")
            report["word_count"] = word_count

            print(f"   Extracted ({word_count} words, {len(extracted_text)} chars)")
            print(f"   Cost: ${cost:.4f}")
            print(f"   Niche: {classification.get('primary_niche', '?')}")

            # Cleanup temp download
            if gdrive_id and local_path and local_path.exists():
                try:
                    local_path.unlink()
                except OSError:
                    pass

        except Exception as e:
            print(f"   Error: {e}")
            supabase.table("files").update({
                "status": "error",
                "error_message": str(e)[:500],
            }).eq("id", file_id).execute()
            report["error"] = str(e)

        return report

    def process_batch(self, limit: int = 10, file_type: str = "all",
                      speed_mode: str = "balanced", local_only: bool = False) -> Dict:
        """Process batch of pending files"""
        files = self.get_pending_files(limit, file_type, local_only=local_only)

        if not files:
            print("No pending files to process")
            return {"processed": 0, "total_cost": 0.0}

        print(f"Processing {len(files)} files...")

        reports = []
        total_cost = 0.0

        for file in files:
            report = self.process_file(file, speed_mode)
            reports.append(report)
            total_cost += report.get("cost_usd", 0.0)

        success_count = sum(1 for r in reports if r["status"] == "success")

        print(f"\nBATCH SUMMARY")
        print(f"   Success: {success_count}/{len(files)}")
        print(f"   Total cost: ${total_cost:.4f}")

        return {
            "processed": len(files),
            "success": success_count,
            "failed": len(files) - success_count,
            "total_cost_usd": round(total_cost, 4),
            "reports": reports,
        }


def main():
    import argparse

    parser = argparse.ArgumentParser(description="Akasha Content Extractor")
    parser.add_argument("--limit", type=int, default=10, help="Max files to process")
    parser.add_argument("--file-type", default="all", choices=["all", "pdf", "video", "audio", "text", "markdown", "json", "image", "document"])
    parser.add_argument("--speed-mode", default="balanced", choices=["conservative", "balanced", "aggressive"])
    parser.add_argument("--local-only", action="store_true", help="Only process local files")
    parser.add_argument("--dry-run", action="store_true", help="Preview only")

    args = parser.parse_args()
    extractor = ContentExtractor()

    if args.dry_run:
        files = extractor.get_pending_files(args.limit, args.file_type, local_only=args.local_only)
        print(f"\nDRY RUN: Would process {len(files)} files\n")
        for i, f in enumerate(files, 1):
            mime = f.get('mime_type', '?')
            etype = get_extract_type(mime)
            size_mb = (f.get('file_size_bytes', 0) or 0) / 1024 / 1024
            print(f"{i}. {f['file_name']} ({etype}, {size_mb:.1f}MB)")
        return

    extractor.process_batch(args.limit, args.file_type, args.speed_mode, args.local_only)


if __name__ == "__main__":
    main()
